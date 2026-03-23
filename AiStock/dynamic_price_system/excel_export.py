#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 导出模块
"""

import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from config import FILE_PATHS
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelExporter:
    """Excel 导出器"""
    
    def __init__(self):
        self.wb = Workbook()
        self.styles = self._init_styles()
    
    def _init_styles(self):
        """初始化样式"""
        return {
            'header': Font(name='微软雅黑', size=11, bold=True, color='FFFFFF'),
            'header_fill': PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid'),
            'normal': Font(name='微软雅黑', size=10),
            'positive': Font(name='微软雅黑', size=10, color='006100'),
            'negative': Font(name='微软雅黑', size=10, color='9C0006'),
            'center': Alignment(horizontal='center', vertical='center'),
        }
    
    def export_dynamic_prices(self, results, filename=None):
        """导出动态价格表"""
        if filename is None:
            filename = FILE_PATHS['output_excel']
        
        ws = self.wb.active
        ws.title = '动态价格'
        
        # 表头
        headers = ['代码', '名称', '板块', '当前价', '入场价', '止损价', '目标价', 
                   '盈亏比', '基本面评分', '宏观系数', '建议']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.styles['header']
            cell.fill = self.styles['header_fill']
            cell.alignment = self.styles['center']
        
        # 数据
        from config import STOCKS_CONFIG
        stock_names = {s['code']: s['name'] for s in STOCKS_CONFIG}
        stock_sectors = {s['code']: s['sector'] for s in STOCKS_CONFIG}
        
        for row, result in enumerate(results, 2):
            code = result['code']
            data = [
                code,
                stock_names.get(code, ''),
                stock_sectors.get(code, ''),
                result['current_price'],
                result['entry_price'],
                result['stop_loss'],
                result['target_price'],
                result['profit_loss_ratio'],
                result['fundamental_score'],
                result['macro_factor'],
                result['recommendation']
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = self.styles['normal']
                cell.alignment = self.styles['center']
                
                # 盈亏比颜色
                if col == 8 and isinstance(value, (int, float)):
                    if value >= 3:
                        cell.font = self.styles['positive']
                    elif value < 1.5:
                        cell.font = self.styles['negative']
        
        # 调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 20)
        
        # 保存
        self.wb.save(filename)
        logger.info(f"✅ Excel 导出成功：{filename}")
        return filename
    
    def export_portfolio_summary(self, summary, filename=None):
        """导出组合摘要"""
        if filename is None:
            filename = FILE_PATHS['output_excel'].replace('.xlsx', '_portfolio.xlsx')
        
        ws = self.wb.create_sheet('组合摘要')
        
        # 基本信息
        info = [
            ['组合总市值', f"¥{summary['total_value']:,.2f}"],
            ['现金', f"¥{summary['cash']:,.2f}"],
            ['总收益率', summary['total_return']],
            ['交易次数', summary['transaction_count']],
            ['更新日期', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ]
        
        for row, (label, value) in enumerate(info, 1):
            ws.cell(row=row, column=1, value=label).font = self.styles['normal']
            ws.cell(row=row, column=2, value=value).font = self.styles['positive']
        
        # 持仓明细
        ws.cell(row=7, column=1, value='持仓明细').font = self.styles['header']
        
        headers = ['代码', '数量', '成本', '现价', '市值', '盈利', '盈利率', '权重']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col, value=header)
            cell.font = self.styles['header']
            cell.fill = self.styles['header_fill']
        
        for row, pos in enumerate(summary['positions'], 9):
            data = [
                pos['code'], pos['quantity'], f"¥{pos['cost']:.2f}",
                f"¥{pos['current_price']:.2f}", f"¥{pos['market_value']:,.2f}",
                f"¥{pos['profit']:,.2f}", pos['profit_pct'], pos['weight']
            ]
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        self.wb.save(filename)
        logger.info(f"✅ 组合摘要导出成功：{filename}")


# 测试
if __name__ == '__main__':
    exporter = ExcelExporter()
    
    # 模拟动态价格结果
    results = [
        {'code': '600938', 'current_price': 42.24, 'entry_price': 40.13, 'stop_loss': 39.20,
         'target_price': 48.50, 'profit_loss_ratio': 3.21, 'fundamental_score': 78.5,
         'macro_factor': 1.03, 'recommendation': '推荐'}
    ]
    
    exporter.export_dynamic_prices(results)