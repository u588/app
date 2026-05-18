"""
数据加载模块
负责XLSX/CSV数据读取、清洗、标准化，以及YAML配置加载
"""

import csv
import os
from typing import List, Dict, Optional
import yaml


class DataLoader:
    """数据加载器，负责读取标的文件和YAML配置文件"""

    def __init__(self, data_path: str, config_dir: str):
        """
        初始化数据加载器

        Args:
            data_path: 标的数据文件路径(.xlsx或.csv)
            config_dir: YAML配置文件目录
        """
        self.data_path = data_path
        self.config_dir = config_dir
        self._targets: List[Dict] = []
        self._industry_chain_config: Dict = {}
        self._relationships_config: Dict = {}
        self._visualization_config: Dict = {}

    @property
    def targets(self) -> List[Dict]:
        """获取标的列表"""
        return self._targets

    @property
    def industry_chain_config(self) -> Dict:
        """获取产业链配置"""
        return self._industry_chain_config

    @property
    def relationships_config(self) -> Dict:
        """获取关系配置"""
        return self._relationships_config

    @property
    def visualization_config(self) -> Dict:
        """获取可视化配置"""
        return self._visualization_config

    def load_all(self) -> 'DataLoader':
        """加载所有数据，返回self以支持链式调用"""
        self._load_data()
        self._load_configs()
        return self

    def _load_data(self):
        """读取标的文件，自动识别XLSX/CSV格式"""
        ext = os.path.splitext(self.data_path)[1].lower()
        if ext == '.xlsx':
            self._load_xlsx()
        elif ext == '.csv':
            self._load_csv()
        else:
            raise ValueError(f"不支持的文件格式: {ext}，仅支持 .xlsx 和 .csv")

    def _load_xlsx(self):
        """读取XLSX标的文件"""
        try:
            import openpyxl
        except ImportError:
            print("[DataLoader] 需要安装openpyxl: pip install openpyxl")
            raise

        wb = openpyxl.load_workbook(self.data_path, data_only=True)
        ws = wb.active

        # 读取表头
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(cell.value.strip().replace('\n', ''))
            else:
                headers.append('')

        # 映射表头到标准化字段名
        header_map = {
            '一级方向': '一级方向',
            '方向优先级': '方向优先级',
            '二级分类': '二级分类',
            '分类优先级': '分类优先级',
            '政策周期': '政策周期',
            '三级赛道': '三级赛道',
            '赛道优先级': '赛道优先级',
            '赛道优先级说明': '赛道优先级说明',
            '赛道独立价值评级': '赛道独立价值评级',
            '独立价值说明': '独立价值说明',
            '名称': '标的名称',
            '代码': '代码',
            '规模分类': '市值规模',
            '投资风格': '投资风格',
            '政策契合度': '政策契合度',
            '投资确定性': '投资确定性',
            '核心业务占比': '核心业务占比',
            '综合评分': '综合评分',
            '标的优先级': '标的优先级',
            '配置建议': '配置建议',
            '优化入选说明': '入选说明',
        }

        targets = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), 2):
            cleaned = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    key = header_map.get(headers[i], headers[i])
                    cleaned[key] = str(val).strip() if val is not None else ''

            target = self._normalize_target(cleaned)
            if target:
                targets.append(target)

        self._targets = targets
        print(f"[DataLoader] 已加载 {len(targets)} 个标的 (XLSX)")

    def _load_csv(self):
        """读取CSV标的文件，处理BOM和编码问题"""
        targets = []
        with open(self.data_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                # 标准化字段名（去除空白和BOM标记）
                cleaned = {}
                for k, v in row.items():
                    key = k.strip().replace('\ufeff', '')
                    cleaned[key] = v.strip() if v else ''
                # 类型转换
                target = self._normalize_target(cleaned)
                if target:
                    targets.append(target)
        self._targets = targets
        print(f"[DataLoader] 已加载 {len(targets)} 个标的 (CSV)")

    def _normalize_target(self, raw: Dict) -> Optional[Dict]:
        """
        标准化单个标的数据

        Args:
            raw: 原始字典数据

        Returns:
            标准化后的标的字典，数据不完整返回None
        """
        name = raw.get('标的名称', '').strip()
        if not name:
            return None

        try:
            policy_score = int(float(raw.get('政策契合度', 0)))
        except (ValueError, TypeError):
            policy_score = 0

        try:
            certainty_score = int(float(raw.get('投资确定性', 0)))
        except (ValueError, TypeError):
            certainty_score = 0

        try:
            score = float(raw.get('综合评分', 0))
        except (ValueError, TypeError):
            score = 0.0

        try:
            target_priority = int(float(raw.get('标的优先级', 0)))
        except (ValueError, TypeError):
            target_priority = 0

        try:
            core_ratio = int(float(raw.get('核心业务占比', 0)))
        except (ValueError, TypeError):
            core_ratio = 0

        try:
            value_rating = int(float(raw.get('赛道独立价值评级', 0)))
        except (ValueError, TypeError):
            value_rating = 0

        code = raw.get('代码', '').strip()
        # 处理代码格式（去掉.0后缀）
        if code.endswith('.0'):
            code = code[:-2]

        # 市值规模映射
        scale_raw = raw.get('市值规模', '小').strip()
        scale_map = {'大盘': '大', '中盘': '中', '小盘': '小'}
        market_cap = scale_map.get(scale_raw, scale_raw)

        return {
            '一级方向': raw.get('一级方向', '').strip(),
            '方向优先级': raw.get('方向优先级', '').strip(),
            '二级分类': raw.get('二级分类', '').strip(),
            '分类优先级': raw.get('分类优先级', '').strip(),
            '政策周期': raw.get('政策周期', '').strip(),
            '三级赛道': raw.get('三级赛道', '').strip(),
            '赛道优先级': raw.get('赛道优先级', '').strip(),
            '赛道优先级说明': raw.get('赛道优先级说明', '').strip(),
            '赛道独立价值评级': value_rating,
            '独立价值说明': raw.get('独立价值说明', '').strip(),
            '标的名称': name,
            '代码': code,
            '市值规模': market_cap,
            '投资风格': raw.get('投资风格', '').strip(),
            '政策契合度': policy_score,
            '投资确定性': certainty_score,
            '核心业务占比': core_ratio,
            '综合评分': score,
            '标的优先级': target_priority,
            '配置建议': raw.get('配置建议', '').strip(),
            '入选说明': raw.get('入选说明', '').strip(),
            # 唯一标识
            'uid': f"{code}_{name}",
        }

    def _load_configs(self):
        """加载所有YAML配置文件"""
        configs = {
            'industry_chain': 'industry_chain.yaml',
            'relationships': 'relationships.yaml',
            'visualization': 'visualization.yaml',
        }
        for attr, filename in configs.items():
            filepath = os.path.join(self.config_dir, filename)
            if os.path.exists(filepath):
                with open(filepath, 'r', encoding='utf-8') as f:
                    config = yaml.safe_load(f)
                setattr(self, f'_{attr}_config', config or {})
                print(f"[DataLoader] 已加载配置: {filename}")
            else:
                print(f"[DataLoader] 警告: 配置文件不存在 {filepath}")

    def get_targets_by_industry(self, industry: str) -> List[Dict]:
        """按一级方向筛选标的"""
        return [t for t in self._targets if t['一级方向'] == industry]

    def get_targets_by_track(self, industry: str, category: str, track: str) -> List[Dict]:
        """按三级赛道筛选标的"""
        return [
            t for t in self._targets
            if t['一级方向'] == industry
            and t['二级分类'] == category
            and t['三级赛道'] == track
        ]

    def get_target_by_name(self, name: str) -> Optional[Dict]:
        """按标的名称查找"""
        for t in self._targets:
            if t['标的名称'] == name:
                return t
        return None

    def get_target_by_code(self, code: str) -> Optional[Dict]:
        """按代码查找"""
        for t in self._targets:
            if t['代码'] == code:
                return t
        return None

    def get_all_industries(self) -> List[str]:
        """获取所有一级方向列表"""
        seen = []
        for t in self._targets:
            if t['一级方向'] not in seen:
                seen.append(t['一级方向'])
        return seen

    def get_tracks_by_industry(self, industry: str) -> Dict[str, List[str]]:
        """获取某一级方向下的二级分类→三级赛道映射"""
        result = {}
        for t in self._targets:
            if t['一级方向'] == industry:
                cat = t['二级分类']
                track = t['三级赛道']
                if cat not in result:
                    result[cat] = []
                if track not in result[cat]:
                    result[cat].append(track)
        return result
